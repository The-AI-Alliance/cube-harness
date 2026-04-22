"""CUBE Tool implementation for DRBench.

DrBenchTool exposes the enterprise adapters (Nextcloud, Mattermost, Email, FileBrowser)
and web search as CUBE @tool_action methods. These become the agent's action_set.

Text extraction helpers are copied from drbench-agent-dev/mcp_servers/nextcloud_server.py
so this module is self-contained (no agent-dev dependency).
"""

from __future__ import annotations

import io
import json
import logging
import os
from pathlib import Path

import requests
from bs4 import BeautifulSoup

from cube.container import Container
from cube.tool import Tool, ToolConfig, tool_action

from drbench.agents.drbench_agent.agent_tools.enterprise.adapters.nextcloud import NextcloudAdapter
from drbench.agents.drbench_agent.agent_tools.enterprise.adapters.mattermost import MattermostAdapter
from drbench.agents.drbench_agent.agent_tools.enterprise.adapters.email import EmailAdapter
from drbench.agents.drbench_agent.agent_tools.enterprise.adapters.filebrowser import FileBrowserAdapter

logger = logging.getLogger(__name__)

_MAX_CONTENT_CHARS = 50_000

# ---------------------------------------------------------------------------
# Text extraction helpers (copied from mcp_servers/nextcloud_server.py)
# ---------------------------------------------------------------------------

_TEXT_EXTENSIONS = {
    ".txt", ".csv", ".tsv", ".md", ".markdown", ".rst", ".html", ".htm",
    ".json", ".xml", ".yaml", ".yml", ".ini", ".cfg", ".conf", ".log",
    ".py", ".js", ".ts", ".java", ".c", ".cpp", ".h", ".sh", ".bash",
    ".sql", ".r", ".rb", ".go", ".rs", ".toml", ".env", ".properties",
}


def _is_text_content(path: str, content_type: str) -> bool:
    if "text" in content_type or content_type in ("application/json", "application/xml"):
        return True
    ext = os.path.splitext(path)[1].lower()
    return ext in _TEXT_EXTENSIONS


def _extract_pdf_text(content: bytes) -> str:
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(content))
        pages = [page.extract_text() for page in reader.pages if page.extract_text()]
        if pages:
            return "\n\n".join(pages)
    except Exception:
        pass
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            pages = [page.extract_text() for page in pdf.pages if page.extract_text()]
            if pages:
                return "\n\n".join(pages)
    except Exception:
        pass
    return ""


def _extract_docx_text(content: bytes) -> str:
    try:
        import docx
        doc = docx.Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs if p.text)
    except Exception:
        return ""


def _extract_xlsx_text(content: bytes) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                lines.append(" | ".join(cells))
        return "\n".join(lines)
    except Exception:
        return ""


def _extract_text_from_bytes(content: bytes, path: str, content_type: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf" or "pdf" in content_type:
        return _extract_pdf_text(content)
    if ext == ".docx" or "wordprocessingml" in content_type:
        return _extract_docx_text(content)
    if ext == ".xlsx" or "spreadsheetml" in content_type:
        return _extract_xlsx_text(content)
    return ""


def _process_download_result(result: dict, path: str) -> dict:
    """Replace raw bytes content with extracted text in an adapter download_file() result."""
    if not result.get("success") or not result.get("content"):
        return result

    content = result["content"]
    content_type = result.get("content_type", "")

    if isinstance(content, bytes):
        if _is_text_content(path, content_type):
            result["content"] = content.decode("utf-8", errors="ignore")[:_MAX_CONTENT_CHARS]
        else:
            extracted = _extract_text_from_bytes(content, path, content_type)
            result["content"] = extracted[:_MAX_CONTENT_CHARS] if extracted else f"[Binary file: {path}]"

    return result


# ---------------------------------------------------------------------------
# Web helpers (copied from mcp_servers/web_server.py)
# ---------------------------------------------------------------------------

_web_session = requests.Session()
_web_session.headers.update({"User-Agent": "Mozilla/5.0 (compatible; DrBench-Agent/1.0; Research-Bot)"})


def _fetch_from_html(content_bytes: bytes) -> str:
    soup = BeautifulSoup(content_bytes, "html.parser")
    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()
    text = soup.get_text()
    lines = (line.strip() for line in text.splitlines())
    chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
    return " ".join(chunk for chunk in chunks if chunk)


def _fetch_from_pdf(content_bytes: bytes) -> str:
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(content_bytes))
        return "\n".join(page.extract_text() for page in reader.pages if page.extract_text()).strip()
    except Exception:
        pass
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(content_bytes)) as pdf:
            return "\n".join(p.extract_text() for p in pdf.pages if p.extract_text()).strip()
    except Exception:
        return ""


# ---------------------------------------------------------------------------
# DrBenchTool
# ---------------------------------------------------------------------------

class DrBenchTool(Tool):
    """
    Environment-side CUBE tool for DRBench.

    Talks directly to container services via HTTP/IMAP using the enterprise
    adapters — no MCP, no subprocess. The agent receives only the action_set
    (OpenAI function-call schemas) and submits Action objects; this class
    executes them.
    """

    def __init__(self, container: Container, username: str, password: str):
        self._container = container
        self._submitted_report: str | None = None

        self._nc = NextcloudAdapter({
            "url": container.get_url(8081),
            "credentials": {"username": username, "password": password},
        })
        self._mm = MattermostAdapter({
            "url": container.get_url(8082),
            "credentials": {"username": username, "password": password},
        })
        self._email = EmailAdapter({
            "url": f"imap://localhost:{container.forward_port(1143)}",
            "host_port": container.forward_port(1143),
            "name": "email_imap",
            "credentials": {"username": username, "password": password},
        })
        self._fb = FileBrowserAdapter({
            "url": container.get_url(8090),
            "credentials": {"username": username, "password": password},
        })

    def reset(self) -> None:
        self._submitted_report = None
        for adapter in (self._nc, self._mm, self._email, self._fb):
            try:
                adapter.discover_capabilities()
            except Exception as e:
                logger.warning(f"discover_capabilities failed for {adapter.__class__.__name__}: {e}")

    # ------------------------------------------------------------------
    # Nextcloud actions
    # ------------------------------------------------------------------

    @tool_action
    def search_nextcloud_files(self, query: str) -> str:
        """Search for files in Nextcloud by keyword match against filenames.

        Tries the Nextcloud Search API; falls back to recursive WebDAV listing
        filtered by filename. Matching is keyword-based against file names — NOT
        full-text search of file contents. Use short, specific keywords likely to
        appear in file names, not full sentences. When results are empty, use
        list_nextcloud_directory("/") to enumerate available files, then
        download_nextcloud_file() on promising candidates.

        Args:
            query: Space-separated keywords to match against filenames
        """
        return json.dumps(self._nc.search(query.split()), default=str)

    @tool_action
    def list_nextcloud_directory(self, path: str = "/") -> str:
        """List files and directories at the given path in Nextcloud.

        Uses WebDAV PROPFIND to enumerate directory contents.

        Args:
            path: Directory path relative to user's files (default: root)
        """
        return json.dumps(self._nc.list_files(path), default=str)

    @tool_action
    def download_nextcloud_file(self, path: str) -> str:
        """Download a file from Nextcloud and return its text content.

        Extracts text from PDFs and common document formats. For text files,
        content is returned directly. Returns extracted text for binary document
        formats (PDF, DOCX, etc.).

        Args:
            path: File path relative to user's files
        """
        result = self._nc.download_file(path)
        return json.dumps(_process_download_result(result, path), default=str)

    # ------------------------------------------------------------------
    # FileBrowser actions
    # ------------------------------------------------------------------

    @tool_action
    def search_filebrowser_files(self, query: str) -> str:
        """Search for files in FileBrowser by keyword match against filenames.

        Recursively lists all files and filters by filename. Matching is
        keyword-based against file names — NOT full-text search of file contents.
        Use short, specific keywords likely to appear in file names, not full
        sentences. When results are empty, use list_filebrowser_directory("/") to
        enumerate available files, then download_filebrowser_file() on promising
        candidates.

        Args:
            query: Space-separated keywords to match against filenames
        """
        return json.dumps(self._fb.search(query.split()), default=str)

    @tool_action
    def list_filebrowser_directory(self, path: str = "/") -> str:
        """List files and directories at the given path in FileBrowser.

        Args:
            path: Directory path (default: root)
        """
        return json.dumps(self._fb.list_files(path), default=str)

    @tool_action
    def download_filebrowser_file(self, path: str) -> str:
        """Download a file from FileBrowser and return its text content.

        For text files, content is returned directly. For binary files (PDF,
        DOCX, XLSX), extracts and returns the text content.

        Args:
            path: File path to download
        """
        result = self._fb.download_file(path)
        return json.dumps(_process_download_result(result, path), default=str)

    # ------------------------------------------------------------------
    # Mattermost actions
    # ------------------------------------------------------------------

    @tool_action
    def search_mattermost(self, query: str) -> str:
        """Search for messages across all Mattermost channels.

        Uses POST /api/v4/posts/search with OR logic across terms. Resolves user
        and channel names for each result. Matching is keyword-based — use short,
        specific words likely to appear in messages, not full sentences. Multiple
        keywords are OR-matched, so prefer fewer, more distinctive terms.

        Args:
            query: Space-separated keywords to search for in message content
        """
        return json.dumps(self._mm.search(query.split()), default=str)

    @tool_action
    def list_mattermost_teams(self) -> str:
        """List all Mattermost teams accessible to the authenticated user."""
        return json.dumps(self._mm.list_files("/"), default=str)

    # ------------------------------------------------------------------
    # Email actions
    # ------------------------------------------------------------------

    @tool_action
    def search_emails(self, query: str) -> str:
        """Search for emails matching the query in INBOX.

        Searches across subject, sender, and body fields using IMAP SEARCH
        with OR logic. Returns up to 20 most recent matches.

        Args:
            query: Search terms to find in emails
        """
        return json.dumps(self._email.search(query.split()), default=str)

    @tool_action
    def get_email(self, folder: str, msg_id: str) -> str:
        """Get the full content of a specific email.

        Args:
            folder: Mailbox folder (e.g., "INBOX")
            msg_id: Message ID within the folder
        """
        result = self._email.download_file(f"{folder}/{msg_id}")
        if result.get("success") and isinstance(result.get("content"), bytes):
            result["content"] = result["content"].decode("utf-8", errors="ignore")
        return json.dumps(result, default=str)

    @tool_action
    def list_email_folders(self) -> str:
        """List available mailbox folders with message counts."""
        return json.dumps(self._email.list_files("/"), default=str)

    # ------------------------------------------------------------------
    # Web actions
    # ------------------------------------------------------------------

    @tool_action
    def web_search(self, query: str, num_results: int = 10) -> str:
        """Search the web using Google via Serper API.

        Returns JSON with organic results (title, link, snippet) plus
        answer box and knowledge graph when available.

        Args:
            query: Search query string
            num_results: Number of results to return (default 10)
        """
        api_key = os.environ.get("SERPER_API_KEY", "")
        if not api_key:
            return json.dumps({"error": "SERPER_API_KEY not set"})

        headers = {"X-API-KEY": api_key, "Content-Type": "application/json"}
        payload = json.dumps({"q": query, "num": num_results})
        try:
            response = requests.post(
                "https://google.serper.dev/search",
                headers=headers,
                data=payload,
                timeout=30,
            )
            response.raise_for_status()
            data = response.json()
            results = [
                {"title": item.get("title"), "link": item.get("link"), "snippet": item.get("snippet")}
                for item in data.get("organic", [])
            ]
            output: dict = {"results": results, "total": len(results)}
            if "answerBox" in data:
                output["answer_box"] = data["answerBox"]
            if "knowledgeGraph" in data:
                output["knowledge_graph"] = data["knowledgeGraph"]
            return json.dumps(output)
        except Exception as e:
            return json.dumps({"error": f"Search failed: {str(e)}"})

    @tool_action
    def fetch_url(self, url: str) -> str:
        """Fetch a URL and extract its text content.

        Supports HTML, PDF, JSON, and plain text. Returns extracted text
        with metadata (title, content_type, content_length).

        Args:
            url: The URL to fetch and extract content from
        """
        try:
            response = _web_session.get(url, timeout=30)
            response.raise_for_status()
            content_type = response.headers.get("content-type", "").lower()
            content_bytes = response.content

            if "text/html" in content_type:
                text = _fetch_from_html(content_bytes)
                soup = BeautifulSoup(content_bytes, "html.parser")
                title_tag = soup.find("title")
                title = title_tag.string.strip()[:200] if title_tag and title_tag.string else "Web Document"
            elif "application/pdf" in content_type or url.lower().endswith(".pdf"):
                text = _fetch_from_pdf(content_bytes)
                title = "PDF Document"
            elif "application/json" in content_type:
                text = json.dumps(json.loads(content_bytes.decode("utf-8", errors="ignore")), indent=2)
                title = "JSON Document"
            else:
                text = content_bytes.decode("utf-8", errors="ignore")
                title = "Web Document"

            return json.dumps({
                "url": url,
                "title": title,
                "content_type": content_type.split(";")[0].strip(),
                "content_length": len(text),
                "content": text[:_MAX_CONTENT_CHARS],
            })
        except Exception as e:
            return json.dumps({"url": url, "error": f"Fetch failed: {str(e)}"})

    # ------------------------------------------------------------------
    # Terminal action
    # ------------------------------------------------------------------

    @tool_action
    def submit_report(self, report_text: str) -> str:
        """Submit the completed research report. Call this when you have gathered
        sufficient information and are ready to present your findings.

        The report should be a comprehensive analysis addressing the research question,
        citing the relevant sources you discovered.

        Args:
            report_text: The complete research report text
        """
        self._submitted_report = report_text
        return json.dumps({"status": "success", "message": "Report submitted successfully."})


# ---------------------------------------------------------------------------
# DrBenchToolConfig
# ---------------------------------------------------------------------------

class DrBenchToolConfig(ToolConfig):
    """Serializable config that instantiates a DrBenchTool from a container."""

    persona_username: str
    persona_password: str

    def make(self, container: Container | None = None) -> DrBenchTool:
        if container is None:
            raise ValueError("DrBench tasks require a container — pass container_backend to TaskConfig.make()")
        return DrBenchTool(container, self.persona_username, self.persona_password)
